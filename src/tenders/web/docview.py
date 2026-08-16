"""Server-side previews for captured documents.

Spreadsheets are the awkward case. No browser renders ``.xls``/``.xlsx``, and a
BOQ or price schedule is often the single most interesting file on a tender —
the thing a reader actually wants to read, not download. So workbooks are parsed
here and handed to the page as plain rows, which the modal and the standalone
viewer both render as an HTML table.

PDFs and images are streamed by ``/doc`` and drawn by the browser instead; this
module only classifies them.
"""

from __future__ import annotations

import csv
import datetime as _dt
import io
from pathlib import Path

PDF_EXT = {".pdf"}
IMAGE_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
SHEET_EXT = {".xls", ".xlsx", ".xlsm", ".csv"}
TEXT_EXT = {".txt", ".csv", ".log", ".md", ".json", ".xml"}

# What /doc?inline=1 may claim a type for. Everything the browser draws itself
# (PDFs, images) has to be listed here or it arrives as a download instead.
INLINE_MEDIA = {
    ".pdf": "application/pdf", ".png": "image/png", ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg", ".gif": "image/gif", ".webp": "image/webp",
    ".bmp": "image/bmp", ".txt": "text/plain", ".csv": "text/plain",
    ".log": "text/plain", ".md": "text/plain", ".json": "text/plain",
    ".xml": "text/plain",
}

# A tender BOQ can declare tens of thousands of rows. Past a few hundred the
# table stops being readable and starts being a way to hang a phone, so the
# preview truncates and points at the download instead.
MAX_ROWS = 400
MAX_COLS = 60
MAX_CELL_CHARS = 400
MAX_TEXT_CHARS = 200_000
# BoQ templates declare hundreds of columns of which a dozen hold data, so the
# real width has to be measured past MAX_COLS before we can honestly say how
# many columns were withheld.
SCAN_COLS = 400


def preview_kind(filename: str) -> str | None:
    """Which preview a filename can get, or None if it is download-only."""
    ext = Path(filename or "").suffix.lower()
    if ext in PDF_EXT:
        return "pdf"
    if ext in IMAGE_EXT:
        return "image"
    if ext in SHEET_EXT:
        return "sheet"
    if ext in TEXT_EXT:
        return "text"
    return None


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        value = value.isoformat(sep=" ") if isinstance(value, _dt.datetime) else value.isoformat()
    text = str(value).strip()
    return text[:MAX_CELL_CHARS] + "…" if len(text) > MAX_CELL_CHARS else text


def _sheet(name: str, rows: list[list[str]], total_rows: int) -> dict:
    """Package one parsed sheet, keeping only the columns that hold data.

    Empty columns are dropped from the *middle* of the sheet, not just the
    right margin, and this is a correctness fix rather than a cosmetic one.

    A GePNIC BoQ is a template with its data scattered across a very sparse
    grid. BOQ_835214.xls declares 243 columns: 0-54 carry the tax/total
    structure, 55-237 are entirely empty, and **238-242 carry the actual line
    items** — the work descriptions, quantities and units, which are the whole
    point of a bill of quantities. Taking a contiguous window of the first
    MAX_COLS therefore did two bad things at once: it spent 55 of its 60
    columns on blanks, squeezing every real heading down to about one character
    wide (a header like "Percentage" rendered as ten stacked letters), and it
    stopped at column 59, so the line items were **silently dropped from the
    preview altogether**. The reader saw a mangled table and had no way to know
    the substantive rows were missing entirely.

    Compacting first fixes both: the 243-column sheet collapses to the ~60
    columns that actually say something, each gets a readable share of the
    width, and nothing real is cut. ``truncated_cols`` counts only columns with
    content that still did not fit, so it stays an honest statement of what the
    reader was denied — blank columns were never information.
    """
    while rows and not any(rows[-1]):
        rows.pop()
    # Column indices holding content anywhere in the sheet, in sheet order.
    keep = sorted({i for row in rows for i, cell in enumerate(row) if cell})
    shown = keep[:MAX_COLS]
    grid = [[(row[i] if i < len(row) else "") for i in shown] for row in rows]
    return {
        "name": name,
        "header": grid[0] if grid else [],
        "rows": grid[1:],
        "cols": len(shown),
        "truncated_rows": max(0, total_rows - MAX_ROWS),
        "truncated_cols": max(0, len(keep) - MAX_COLS),
    }


def _read_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out = []
        for ws in wb.worksheets:
            rows, seen = [], 0
            for raw in ws.iter_rows(values_only=True):
                seen += 1
                if len(rows) < MAX_ROWS:
                    rows.append([_cell(v) for v in raw[:SCAN_COLS]])
            out.append(_sheet(ws.title, rows, seen))
        return out
    finally:
        wb.close()


def _open_xls(path: Path):
    import xlrd
    import xlrd.book

    try:
        return xlrd.open_workbook(path, formatting_info=False)
    except ValueError:
        # A defined name whose formula references a deleted sheet makes xlrd
        # abort the whole workbook. Those names are irrelevant to a preview, so
        # retry with name evaluation disabled rather than lose the file.
        original = xlrd.book.evaluate_name_formula
        xlrd.book.evaluate_name_formula = lambda *a, **k: None
        try:
            return xlrd.open_workbook(path, formatting_info=False)
        finally:
            xlrd.book.evaluate_name_formula = original


def _read_xls(path: Path) -> list[dict]:
    import xlrd

    book = _open_xls(path)
    try:
        out = []
        for ws in book.sheets():
            rows = []
            for r in range(min(ws.nrows, MAX_ROWS)):
                row = []
                for c in range(min(ws.ncols, SCAN_COLS)):
                    value = ws.cell_value(r, c)
                    if ws.cell_type(r, c) == xlrd.XL_CELL_DATE:
                        try:
                            value = _dt.datetime(*xlrd.xldate_as_tuple(value, book.datemode))
                        except ValueError:  # XLDateError included; keep the raw serial
                            pass
                    row.append(_cell(value))
                rows.append(row)
            out.append(_sheet(ws.name, rows, ws.nrows))
        return out
    finally:
        book.release_resources()


def _decode(path: Path) -> str:
    raw = path.read_bytes()[: MAX_TEXT_CHARS * 2]
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "replace")


def _read_csv(path: Path) -> list[dict]:
    text = _decode(path)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows, seen = [], 0
    for raw in csv.reader(io.StringIO(text), dialect):
        seen += 1
        if len(rows) < MAX_ROWS:
            rows.append([_cell(v) for v in raw[:SCAN_COLS]])
    return [_sheet("Sheet 1", rows, seen)]


_READERS = {".xlsx": _read_xlsx, ".xlsm": _read_xlsx, ".xls": _read_xls, ".csv": _read_csv}


def build_preview(path: Path, filename: str) -> dict:
    """Everything a template needs to render one document's preview.

    Sheet parsing can fail on a corrupt or misnamed capture; that is a preview
    problem, not a page problem, so it degrades to an error note beside a
    working download link rather than a 500."""
    kind = preview_kind(filename)
    view: dict = {"kind": kind, "sheets": [], "text": "", "error": ""}
    if kind == "sheet":
        reader = _READERS.get(Path(filename).suffix.lower())
        try:
            view["sheets"] = reader(path) if reader else []
        except Exception as exc:  # noqa: BLE001 - any parse failure degrades the same way
            view["sheets"] = []
            view["error"] = f"This spreadsheet could not be read ({type(exc).__name__})."
        if not view["error"] and not any(s["rows"] or s["header"] for s in view["sheets"]):
            view["error"] = "This spreadsheet is empty."
        # A .csv that isn't really tabular still reads fine as plain text.
        if view["error"] and Path(filename).suffix.lower() == ".csv":
            view["kind"] = kind = "text"
    if kind == "text":
        try:
            text = _decode(path)
        except OSError as exc:
            view["text"], view["error"] = "", f"This file could not be read ({exc.strerror})."
        else:
            view["truncated_text"] = len(text) > MAX_TEXT_CHARS
            view["text"] = text[:MAX_TEXT_CHARS]
            view["error"] = ""
    return view
