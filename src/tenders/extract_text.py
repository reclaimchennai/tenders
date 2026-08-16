"""Extract searchable text from captured documents.

Strategy per file type:
* PDF: read the embedded text layer first (pdfplumber). If a page yields too few
  characters it is treated as scanned and OCR'd (pdf2image -> Tesseract).
* Spreadsheets: openpyxl for the OOXML family, xlrd for legacy BIFF .xls.
* Plain text / CSV: decoded as-is.
Everything else is recorded as ``unsupported:<ext>``. Extraction is incremental:
only captured documents without a doc_text row are processed.

Every document processed gets a doc_text row **even when it yields no text** —
that row is what keeps the file out of the next pass's queue. Without it a
missing file, an unsupported type or a hard parse failure is re-selected on
every cycle forever, and (because the queue is a bare ``LIMIT``) starves the
documents behind it.
"""

from __future__ import annotations

import contextlib
import logging
import signal
import threading
import time
from pathlib import Path

from .config import load_config
from .db import init_db
from .index_fts import (
    begin_immediate,
    commit,
    ensure_docs_fts_aligned,
    index_document,
    open_writer,
)
from .util import now_iso

log = logging.getLogger("extract")

# The live scraper holds write locks for seconds at a time; db.connect() sets no
# busy_timeout, so without this an extraction pass dies on "database is locked"
# mid-batch and the whole backlog silently stops advancing.
BUSY_TIMEOUT_MS = 30_000

# Wall-clock ceiling per document. max_ocr_pages bounds page count but not the
# time a single pathological page can burn inside poppler/tesseract, and one
# such file must not stall a multi-hour catch-up.
DEFAULT_DOC_TIMEOUT_S = 1200

# Pages rendered per pdf2image call. Rendering page-at-a-time re-parses the
# whole PDF once per page (O(n) poppler startups for an n-page scan); batching
# amortises that, while a small batch keeps peak image RAM bounded.
OCR_BATCH_PAGES = 8

# Pages read per pdfplumber document handle (see _extract_pdf).
PDF_PAGE_CHUNK = 20

# Time budget for the OCR phase of a single PDF. max_ocr_pages bounds how many
# pages are OCR'd but not how long each takes, and an A0 drawing can be a minute
# on its own. Exceeding this stops OCR and keeps whatever the text layer and the
# completed pages produced, rather than letting the hard per-document timeout
# throw all of it away.
OCR_BUDGET_S = 600

# Guard against a single corrupt/generated file producing runaway text.
MAX_TEXT_CHARS = 8_000_000

# Documents handed to one worker pool before it is torn down and recreated.
# Tearing the pool down periodically is what recycles workers; the obvious
# alternative, ProcessPoolExecutor(max_tasks_per_child=...), silently switches
# the start method to "spawn" and deadlocked this pass with no output at all.
SLICE_SIZE = 100

# Address-space ceiling per worker. A single 17 MB scanned PDF drives pdfplumber
# past 2 GB, and several workers doing that at once would take the whole box
# down — including the live scraper, whose captures are unrepeatable. Exceeding
# the cap raises MemoryError inside one worker, which is recorded as an ordinary
# per-document 'error'.
WORKER_ADDRESS_SPACE_BYTES = 2_500_000_000

# Methods that represent "we could not get text this time" and are therefore
# safe to re-attempt when the caller asks for a retry pass.
RETRYABLE_METHODS = ("error", "missing", "timeout")


class ExtractTimeout(Exception):
    pass


@contextlib.contextmanager
def _time_limit(seconds: float):
    """Abort the enclosed work after ``seconds`` using SIGALRM.

    Only usable on the main thread; elsewhere (and when disabled) this is a
    no-op rather than an error, since a missing timeout is far better than
    refusing to extract at all.
    """
    if seconds <= 0 or threading.current_thread() is not threading.main_thread():
        yield
        return

    def _fire(_signum, _frame):
        raise ExtractTimeout(f"exceeded {seconds:.0f}s")

    previous = signal.signal(signal.SIGALRM, _fire)
    signal.setitimer(signal.ITIMER_REAL, seconds)
    try:
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous)


def _sniff(path: Path) -> str:
    """Container format from magic bytes.

    GePNIC filenames lie routinely — BOQ files named .xls are often real .xlsx
    and vice versa — so the parser is chosen by content, not by suffix.
    """
    try:
        with open(path, "rb") as fh:
            head = fh.read(8)
    except OSError:
        return ""
    if head[:4] == b"PK\x03\x04":
        return "zip"
    if head[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "ole2"
    if head[:4] == b"%PDF":
        return "pdf"
    if head[:4] == b"Rar!":
        return "rar"
    return ""


def _contiguous_batches(pages: list[int], size: int):
    """Group sorted page indexes into runs of consecutive pages, capped at
    ``size`` so each pdf2image call renders one contiguous range."""
    batch: list[int] = []
    for p in pages:
        if batch and (p != batch[-1] + 1 or len(batch) >= size):
            yield batch
            batch = []
        batch.append(p)
    if batch:
        yield batch


def _extract_pdf(path: Path, cfg) -> tuple[str, str]:
    import pdfplumber

    ocr_cfg = cfg.ocr
    min_chars = int(ocr_cfg["min_chars_per_page"])
    max_ocr = int(ocr_cfg["max_ocr_pages"])
    ocr_enabled = bool(ocr_cfg["enabled"])

    with pdfplumber.open(path) as pdf:
        n_pages = len(pdf.pages)
    texts: list[str] = [""] * n_pages
    scanned_pages: list[int] = []

    # Reopened per chunk rather than read in one pass: pdfminer caches every
    # indirect object it resolves for the lifetime of the document, so an
    # 828-page tender drawing set climbs past 2 GB before it finishes. Closing
    # the document every PDF_PAGE_CHUNK pages holds that near 120 MB.
    for start in range(0, n_pages, PDF_PAGE_CHUNK):
        wanted = list(range(start + 1, min(start + PDF_PAGE_CHUNK, n_pages) + 1))
        with pdfplumber.open(path, pages=wanted) as pdf:
            for offset, page in enumerate(pdf.pages):
                i = (page.page_number - 1) if page.page_number else (start + offset)
                try:
                    t = page.extract_text() or ""
                except Exception as exc:  # noqa: BLE001
                    # One unparseable page must not cost us the other 827.
                    log.debug("page %d of %s unreadable: %s", i, path.name, exc)
                    t = ""
                finally:
                    with contextlib.suppress(Exception):
                        page.flush_cache()
                if 0 <= i < n_pages:
                    texts[i] = t
                if len(t.strip()) < min_chars:
                    scanned_pages.append(i)
    scanned_pages.sort()

    method = "pdf_layer"
    if scanned_pages and ocr_enabled and _tesseract_ok():
        try:
            import tempfile

            from pdf2image import convert_from_path
            from PIL import Image
            import pytesseract

            done = 0
            deadline = time.monotonic() + OCR_BUDGET_S
            for batch in _contiguous_batches(scanned_pages[:max_ocr], OCR_BATCH_PAGES):
                if time.monotonic() > deadline:
                    log.warning("OCR budget spent on %s after %d page(s)",
                                path.name, done)
                    break
                # Rendered to disk rather than memory: tender drawings are
                # routinely A0/A1, and one such page at 200 dpi is ~180 MB
                # decoded — a batch of them held in RAM was enough to push this
                # box towards the OOM killer.
                with tempfile.TemporaryDirectory(prefix="tender-ocr-") as tmp:
                    pages = convert_from_path(
                        path, first_page=batch[0] + 1, last_page=batch[-1] + 1,
                        dpi=200, output_folder=tmp, paths_only=True,
                    )
                    for offset, image_path in enumerate(pages):
                        page_no = batch[0] + offset
                        if page_no >= len(texts):
                            continue
                        with Image.open(image_path) as image:
                            texts[page_no] = pytesseract.image_to_string(image)
                        done += 1
            # Only claim 'ocr' if OCR actually contributed; a failed attempt
            # left the text layer in place and should be labelled as such.
            if done:
                method = "ocr"
        except ExtractTimeout:
            raise
        except Exception as exc:  # noqa: BLE001
            log.warning("OCR failed for %s: %s", path.name, exc)
    return "\n".join(texts).strip(), method


def _extract_xlsx(path: Path) -> tuple[str, str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[str] = []
    try:
        for ws in wb.worksheets:
            out.append(f"# {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    out.append("\t".join(cells))
    finally:
        wb.close()
    return "\n".join(out).strip(), "xlsx"


def _extract_xls(path: Path) -> tuple[str, str]:
    import xlrd

    book = xlrd.open_workbook(path)
    out: list[str] = []
    for sh in book.sheets():
        out.append(f"# {sh.name}")
        for r in range(sh.nrows):
            cells = [str(c) for c in sh.row_values(r) if str(c).strip()]
            if cells:
                out.append("\t".join(cells))
    return "\n".join(out).strip(), "xls"


def _extract_plain(path: Path) -> tuple[str, str]:
    data = path.read_bytes()[:MAX_TEXT_CHARS]
    for encoding in ("utf-8", "utf-16", "cp1252"):
        try:
            return data.decode(encoding).strip(), "text"
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace").strip(), "text"


def _tesseract_ok() -> bool:
    from .captcha import tesseract_available

    return tesseract_available()


def extract_one(path: Path, cfg) -> tuple[str, str]:
    """Return ``(text, method)`` for one file. Never raises for content reasons;
    a parse failure becomes ``('', 'error')`` so the caller can record it."""
    ext = path.suffix.lower()
    kind = _sniff(path)
    try:
        if ext == ".pdf" or kind == "pdf":
            return _extract_pdf(path, cfg)
        if ext in (".xlsx", ".xlsm") or (ext == ".xls" and kind == "zip"):
            return _extract_xlsx(path)
        if ext == ".xls":
            try:
                return _extract_xls(path)
            except Exception:
                # Some "legacy" BOQ files are OOXML or SpreadsheetML with an
                # .xls suffix that also fails the magic-byte sniff.
                return _extract_xlsx(path)
        if ext in (".txt", ".csv", ".tsv", ".xml", ".json"):
            return _extract_plain(path)
        if ext == ".zip":
            # download_docs.explode_zip already unpacks every bundle member into
            # its own captured document row, so the members are indexed
            # individually. Re-reading the archive here would only duplicate
            # them; the row exists purely to retire the zip from the queue.
            return "", "zip-bundle"
        if ext == ".rar" or kind == "rar":
            # No unrar/unar on the host and no pure-Python RAR5 decoder in the
            # dependency set. Recorded explicitly so these three files show up
            # as a known gap instead of being re-tried every cycle.
            return "", "unsupported:rar"
    except ExtractTimeout:
        raise
    except Exception as exc:  # noqa: BLE001
        log.warning("extract error %s: %s", path.name, exc)
        return "", "error"
    return "", f"unsupported:{ext.lstrip('.') or 'noext'}"


def _extract_for_row(rel: str, root: Path, cfg, doc_timeout: float) -> tuple[str, str]:
    """Resolve one queued document to ``(text, method)``, absorbing every
    failure mode into a recordable method so the caller always writes a row."""
    path = root / rel
    if not path.exists():
        log.warning("captured file absent on disk: %s", rel)
        return "", "missing"
    try:
        with _time_limit(doc_timeout):
            return extract_one(path, cfg)
    except ExtractTimeout as exc:
        log.warning("extract timed out for %s: %s", rel, exc)
        return "", "timeout"
    except Exception as exc:  # noqa: BLE001
        # Per-document isolation: nothing a single file does may end the pass,
        # but it must leave a visible trace.
        log.warning("extract failed for %s: %s", rel, exc)
        return "", "error"


def _init_worker() -> None:
    import os
    import resource

    # Tesseract defaults to OpenMP across all cores, so N workers each spawn N
    # threads and spend more time contending than recognising — and this box
    # also runs the live scraper. One thread per worker keeps the parallelism
    # where we control it.
    os.environ.setdefault("OMP_THREAD_LIMIT", "1")

    _soft, hard = resource.getrlimit(resource.RLIMIT_AS)
    cap = WORKER_ADDRESS_SPACE_BYTES
    if hard != resource.RLIM_INFINITY:
        cap = min(cap, hard)
    with contextlib.suppress(ValueError, OSError):
        resource.setrlimit(resource.RLIMIT_AS, (cap, hard))


def _worker(args: tuple[int, str, str, float]) -> tuple[int, str, str]:
    doc_id, rel, root, doc_timeout = args
    text, method = _extract_for_row(rel, Path(root), load_config(), doc_timeout)
    return doc_id, text, method


def _extract_slice(slice_rows, root: Path, cfg, doc_timeout: float, workers: int):
    """Yield ``(doc_id, text, method)`` for a slice, in parallel when asked.

    Parsing and OCR are CPU-bound and wholly independent per file, so a process
    pool is the one place real speed is available. A fresh pool per slice keeps
    a worker crash (poppler and tesseract are native code) contained: the slice
    falls back to in-process extraction instead of losing the whole run.
    """
    if workers <= 1:
        for row in slice_rows:
            text, method = _extract_for_row(row["stored_path"], root, cfg, doc_timeout)
            yield row["id"], text, method
        return

    from concurrent.futures import ProcessPoolExecutor, as_completed
    from concurrent.futures.process import BrokenProcessPool

    payload = [(r["id"], r["stored_path"], str(root), doc_timeout) for r in slice_rows]
    done: set[int] = set()
    try:
        with ProcessPoolExecutor(max_workers=workers, initializer=_init_worker) as pool:
            # as_completed, not map: map yields in submission order, so a single
            # 800-page document at the head of a slice would hold back every
            # result behind it — and with them the commits that make a
            # multi-hour run resumable.
            futures = [pool.submit(_worker, item) for item in payload]
            for fut in as_completed(futures):
                doc_id, text, method = fut.result()
                done.add(doc_id)
                yield doc_id, text, method
    except BrokenProcessPool as exc:
        log.warning("worker pool died (%s); finishing slice in-process", exc)
        for row in slice_rows:
            if row["id"] in done:
                continue
            text, method = _extract_for_row(row["stored_path"], root, cfg, doc_timeout)
            yield row["id"], text, method


def run_extract(db_path=None, *, limit: int = 0, retry: bool = False,
                max_seconds: float = 0, workers: int = 1) -> dict:
    """Extract text for captured documents that have none yet.

    limit        maximum documents this pass; 0 (the default) means the whole
                 backlog. The old default of 200 structurally starved the
                 pipeline: run_forward() calls this once per cycle after a
                 multi-hour capture loop, so the queue grew thousands of
                 documents per cycle while draining only a couple hundred.
    retry        also re-attempt documents previously recorded as error/
                 missing/timeout (not unsupported ones, which will not change).

    A document re-published on the portal is always re-extracted: store_document
    drops its doc_text row on a sha256 change, and the extracted_at/downloaded_at
    comparison below catches any other path that replaces the bytes in place, so
    text describing a superseded version can never linger in the index.
    max_seconds  optional wall-clock budget for the whole pass; 0 = unbounded.
    workers      parallel extraction processes; 1 (the default) keeps the
                 daemon's behaviour unchanged, higher values are for catch-up
                 runs where OCR dominates.
    """
    from tqdm import tqdm

    cfg = load_config()
    db_path = db_path or cfg.db_path
    init_db(db_path)
    conn = open_writer(db_path)
    root = cfg.db_path.parent.parent
    doc_timeout = float(cfg.ocr.get("per_doc_timeout_s", DEFAULT_DOC_TIMEOUT_S))

    started = time.monotonic()
    processed = ok = 0
    methods: dict[str, int] = {}
    try:
        where_retry = ""
        params: list = []
        if retry:
            placeholders = ",".join("?" * len(RETRYABLE_METHODS))
            where_retry = f" OR t.method IN ({placeholders})"
            params.extend(RETRYABLE_METHODS)
        sql = (
            "SELECT d.id, d.stored_path, d.tender_id, d.filename FROM documents d "
            "LEFT JOIN doc_text t ON t.document_id = d.id "
            "WHERE d.status='captured' AND d.stored_path IS NOT NULL "
            f"  AND (t.document_id IS NULL OR t.extracted_at < d.downloaded_at"
            f"       {where_retry}) "
            "ORDER BY d.id"
        )
        if limit:
            sql += " LIMIT ?"
            params.append(limit)
        rows = conn.execute(sql, params).fetchall()
        log.info("extract: %d document(s) queued", len(rows))
        ensure_docs_fts_aligned(conn)
        meta = {r["id"]: (r["tender_id"], r["filename"]) for r in rows}

        bar = tqdm(total=len(rows), unit="doc", ncols=90, disable=not _isatty(),
                   bar_format="{l_bar}{bar}| {n_fmt}/{total_fmt} [{elapsed}<{remaining}]")
        out_of_time = False
        for start in range(0, len(rows), SLICE_SIZE):
            if out_of_time:
                break
            chunk = rows[start:start + SLICE_SIZE]
            for doc_id, text, method in _extract_slice(chunk, root, cfg, doc_timeout,
                                                       workers):
                if len(text) > MAX_TEXT_CHARS:
                    text = text[:MAX_TEXT_CHARS]
                # Read before overwriting: with an external-content index this
                # is the last moment the previously indexed text exists, and
                # FTS5 needs it back verbatim to unpick the old terms.
                prev = conn.execute(
                    "SELECT text FROM doc_text WHERE document_id=? AND text<>''",
                    (doc_id,),
                ).fetchone()
                conn.execute(
                    "INSERT OR REPLACE INTO doc_text (document_id, text, method, "
                    "char_count, extracted_at) VALUES (?, ?, ?, ?, ?)",
                    (doc_id, text, method, len(text), now_iso()),
                )
                tender_id, filename = meta.get(doc_id, (None, None))
                index_document(conn, doc_id, tender_id, filename, text,
                               old_text=prev[0] if prev else None)
                # Committed per document rather than per batch: a batch would
                # hold the single write lock across minutes of OCR and block the
                # live scraper. WAL + synchronous=NORMAL makes these cheap.
                commit(conn)

                processed += 1
                bar.update(1)
                methods[method] = methods.get(method, 0) + 1
                if text:
                    ok += 1
                if processed % 100 == 0:
                    log.info("extract: %d/%d done (%d with text, %.0fs elapsed)",
                             processed, len(rows), ok, time.monotonic() - started)
            if max_seconds and time.monotonic() - started > max_seconds:
                log.info("extract: time budget reached after %d document(s)", processed)
                out_of_time = True
        bar.close()
        commit(conn)
    finally:
        conn.close()

    result = {"processed": processed, "with_text": ok, "methods": methods,
              "elapsed_s": round(time.monotonic() - started, 1)}
    log.info("extract done: %s", result)
    return result


def _isatty() -> bool:
    import sys

    return bool(getattr(sys.stderr, "isatty", lambda: False)())
